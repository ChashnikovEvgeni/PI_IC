from openpyxl.styles import Side, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter

from IC.models import Indicator, Department, Critical_service


def set_Degree_of_compliance(indicator):
    degree_of_compliance = indicator.actual_indicator / indicator.Plan * 100
    indicator.Degree_of_compliance = degree_of_compliance



def set_department_indicator(department):
    department.department_indicator = 0
    for i in Indicator.objects.filter(department=department):
        department.department_indicator +=  i.Significance_of_indicator * i.actual_indicator / i.Plan * 100
    department.PPRTD_indicator = department.department_indicator*department.PPRTD_weight
    department.PFVIR_indicator = department.department_indicator*department.PFVIR_weight



def set_indicators_PPRTD_PFVIR(service):
    service.indicator_PPRTD = 0
    service.indicator_PFVIR = 0
    service.sum_weight_PPRTD = 0
    service.sum_weight_PFVIR = 0

    for d in Department.objects.filter(service=service):
        service.indicator_PPRTD += d.PPRTD_indicator
        service.indicator_PFVIR += d.PFVIR_indicator
        service.sum_weight_PPRTD += d.PPRTD_weight
        service.sum_weight_PFVIR += d.PFVIR_weight

def set_operating_time(critical_service):
    critical_service.Operating_time_plan = critical_service.working_days_period * critical_service.working_mode_hours
    critical_service.Completion_rate = round(critical_service.Operating_time_actual / critical_service.Operating_time_plan, 3)



def print_settlement_form(request, workbook, index):
    ind_queryset = Indicator.objects.filter(department__in=request.user.profile.access.all())
    worksheet = workbook.create_sheet(title='Расчётная форма', index=index)

    cell = worksheet.cell(row=1, column=1)
    cell.value = 'Единый показатель оценки работы СИУС "Готовность локальных ИУС'
    cell.border = Border(
        bottom = Side(border_style='medium', color ='000000'),
        top=Side(border_style='medium', color='000000'),
        left = Side(border_style='medium', color='000000'),
        right = Side(border_style='medium', color='000000'),
    )
    columns = [
        '#',
        'Название отдела/группы',
        'Название показателя',
        'Единицы измерения показателя',
        'Комментарий',
        'Целевой показаетль',
        'Вес показателя',
        'План',
        'Факт',
        'Значение показателя за отчётный период',
        'C учетом веса показателя по отделу',
    ]
    row_num = 2
    colors = ['A6CAF0', 'FFFFFF']
    color = colors[0]
    dep_title = ''
    wide_columns = [2,3,4,5]
    number = 0
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(
            bottom=Side(border_style='medium', color='000000'),
            top=Side(border_style='medium', color='000000'),
            left=Side(border_style='medium', color='000000'),
            right=Side(border_style='medium', color='000000'),
        )
        cell.alignment = Alignment(horizontal='center')

    for ind in ind_queryset:
        row_num += 1
        number +=1
        worksheet.row_dimensions[row_num].height = 60
        cell = worksheet.cell(row=row_num, column=11)
        if dep_title != ind.department.title:
            dep_title = ind.department.title
            if color == colors[1]:
                color=colors[0]
            else:
                color = colors[1]
            cell.value = ind.department.department_indicator
            cell.border = Border(
                top=Side(border_style='medium', color='000000'),
                left=Side(border_style='medium', color='000000'),
                right=Side(border_style='medium', color='000000'),
            )
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type='solid'
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            column_letter = get_column_letter(11)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 40
        else:
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type='solid'
            )
            cell.border = Border(
                bottom=Side(border_style='medium', color='FFFFFF'),
                top=Side(border_style='medium', color=color),
                left=Side(border_style='medium', color='000000'),
                right=Side(border_style='medium', color='000000'),
            )
        if row_num-2 == ind_queryset.count():
            cell.border = Border(
                bottom=Side(border_style='medium', color='000000'),
                top=Side(border_style='medium', color=color),
                left=Side(border_style='medium', color='000000'),
                right=Side(border_style='medium', color='000000'),
            )

        row = [
            number,
            ind.department.title,
            ind.title,
            ind.units,
            ind.comment,
            ind.target_indicator,
            ind.Significance_of_indicator,
            ind.Plan,
            ind.actual_indicator,
            ind.Degree_of_compliance,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            if col_num in wide_columns:
                column_dimensions.width = 70
                cell.alignment = Alignment(wrapText=True, vertical='top')
            elif col_num==10:
                column_dimensions.width = 50
                cell.alignment = Alignment( horizontal="center", vertical="center")
            else:
                column_dimensions.width = 30
                cell.alignment = Alignment( horizontal="center", vertical="center")
            cell.border = Border(
                bottom=Side(border_style='medium', color='000000'),
                top=Side(border_style='medium', color='000000'),
                left=Side(border_style='medium', color='000000'),
                right=Side(border_style='medium', color='000000'),
            )
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type='solid'
            )



def print_report_departments(request, workbook, index):
    access_departments = request.user.profile.access.all()
    departments_names = []
    for a in access_departments:
        departments_names.append(a.title)
    dep_queryset = Department.objects.filter(title__in=departments_names)
    worksheet = workbook.create_sheet(title='Отчёт отделы', index=index)
    columns = [
        '#',
        'Название отдела/группы',
        'Показатель',
        'Вес, приоритет передача режимно- технологических данных',
        'Показатель',
        'Вес, приоритет функционирование ВИР',
        'Показатель',
    ]
    row_num = 1
    number = 0
    colors = ['A6CAF0', 'A9D08E']
    summs = [0,0,0,0]
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(
            bottom=Side(border_style='medium', color='000000'),
            top=Side(border_style='medium', color='000000'),
            left=Side(border_style='medium', color='000000'),
            right=Side(border_style='medium', color='000000'),
        )
        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    worksheet.row_dimensions[1].height = 30
    for department in dep_queryset:
        row_num += 1
        number += 1
        worksheet.row_dimensions[row_num].height = 30
        row = [
            number,
            department.title,
            department.department_indicator,
            department.PPRTD_weight,
            department. PPRTD_indicator,
            department.PFVIR_weight,
            department.PFVIR_indicator,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 50
            cell.border = Border(
                bottom=Side(border_style='medium', color='000000'),
                top=Side(border_style='medium', color='000000'),
                left=Side(border_style='medium', color='000000'),
                right=Side(border_style='medium', color='000000'),
            )
            if col_num > 3 and col_num < 6:
                cell.fill = PatternFill(
                    start_color=colors[1],
                    end_color=colors[1],
                    fill_type='solid'
                )
            elif col_num > 5:
                cell.fill = PatternFill(
                    start_color=colors[0],
                    end_color=colors[0],
                    fill_type='solid'
                )
            if col_num != 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top')


    cells = []

    cell = worksheet.cell(row=row_num + 1, column=4)
    cells.append(cell)
    cell.fill = PatternFill(
        start_color=colors[1],
        end_color=colors[1],
        fill_type='solid'
    )
    cell.value = dep_queryset[0].service.sum_weight_PPRTD
    cell = worksheet.cell(row=row_num + 1, column=5)
    cells.append(cell)
    cell.fill = PatternFill(
        start_color=colors[1],
        end_color=colors[1],
        fill_type='solid'
    )
    cell.value = dep_queryset[0].service.indicator_PPRTD

    cell = worksheet.cell(row=row_num + 1, column=6)
    cells.append(cell)
    cell.fill = PatternFill(
        start_color=colors[0],
        end_color=colors[0],
        fill_type='solid'
    )
    cell.value = dep_queryset[0].service.sum_weight_PFVIR


    cell = worksheet.cell(row=row_num + 1, column=7)
    cells.append(cell)
    cell.fill = PatternFill(
        start_color=colors[0],
        end_color=colors[0],
        fill_type='solid'
    )
    cell.value = dep_queryset[0].service.indicator_PFVIR

    cell = worksheet.cell(row=row_num+2, column=2)
    cells.append(cell)
    cell.value = 'Целевой показатель СИУС "Готовность локальных ИУС",%'
    cell = worksheet.cell(row=row_num + 2, column=3)
    cells.append(cell)
    cell.value = dep_queryset[0].service.target_indicator

    cell = worksheet.cell(row=row_num + 3, column=2)
    cells.append(cell)
    cell.value = 'Значение при достижении целевых показателей отделами и группами СИУС:'
    cell = worksheet.cell(row=row_num + 3, column=5)
    cells.append(cell)
    cell.value = dep_queryset[0].service.value_when_reached

    cell = worksheet.cell(row=row_num + 3, column=7)
    cells.append(cell)
    cell.value = dep_queryset[0].service.value_when_reached

    for c in cells:
        c.border = Border(
            bottom=Side(border_style='medium', color='000000'),
            top=Side(border_style='medium', color='000000'),
            left=Side(border_style='medium', color='000000'),
            right=Side(border_style='medium', color='000000'),
        )
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


def print_report_CS(workbook, index):
    cs_queryset = Critical_service.objects.all()
    worksheet = workbook.create_sheet(title='Критические сервисы службы', index=index)
    columns = [
        '#',
        'Название',
        'Рабочих часов в день',
        'Рабочих дней в неделю',
        'Рабочих дней в отчётном периоде',
        'План, ч',
        'Факт, ч',
        'Выполнение',
        'Принадлежность',
    ]
    row_num = 1
    number = 0
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(
            bottom=Side(border_style='medium', color='000000'),
            top=Side(border_style='medium', color='000000'),
            left=Side(border_style='medium', color='000000'),
            right=Side(border_style='medium', color='000000'),
        )
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    for CS in cs_queryset:
        row_num += 1
        number += 1
        worksheet.row_dimensions[row_num].height = 60
        row = [
            number,
            CS.title,
            CS.working_mode_hours,
            CS.working_mode_days,
            CS.working_days_period,
            CS.Operating_time_plan,
            CS.Operating_time_actual,
            CS.Completion_rate,
            CS.Service_ownership,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 30
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(border_style='medium', color='000000'),
                top=Side(border_style='medium', color='000000'),
                left=Side(border_style='medium', color='000000'),
                right=Side(border_style='medium', color='000000'),
            )