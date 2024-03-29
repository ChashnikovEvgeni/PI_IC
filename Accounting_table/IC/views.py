from datetime import datetime

from django.contrib.auth import logout, login
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.views import LoginView
from django.db.models import Q
from django.http import HttpResponse, Http404, HttpRequest
from django.shortcuts import render, redirect  # шаблонизатор фреймворка Django
from django.views.generic import CreateView, ListView
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from requests import Response
from rest_framework import renderers, permissions, status
from rest_framework.decorators import action, api_view, permission_classes, parser_classes
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.reverse import reverse_lazy
from rest_framework.viewsets import ModelViewSet
from rest_framework.pagination import PageNumberPagination
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from .forms import *
from .logic import print_settlement_form, print_report_departments, print_report_CS
from .models import *
from .permissions import IsOwnerOrStaffOrReadOnly, IsAccess, IsCreator, AccessIndicator
from .serializers import *
from .utils import *

@login_required(login_url='login/')
def index(request):
    return render(request, 'IC/index.html')


# Набор представлений службы
class ServiceViewSet(ModelViewSet):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer
    permission_classes =  (IsAccess, )

@api_view(['GET', 'POST'])
@parser_classes([MultiPartParser])
@permission_classes([IsCreator])
def forms_service(request, service_id=None):
    if service_id is None:
        service = None
        template = 'IC/add_object.html'
        context = {
            'url_name': 'add_service',
            'title': 'Добавление службы',
        }
    else:
        service = Service.objects.get(pk=service_id)
        template = 'IC/change_object.html'
        context = {
            'url_name': 'change_service',
            'title': 'Редактирование службы',
        }
    if request.method == 'POST':
        form = ServiceForm(request.POST, request.FILES, instance=service)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = ServiceForm(instance=service)
        context['form'] = form
    return render(request, template, context)


# Набор представлений отдела/группы



class DepartmentViewSet(ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = (IsAccess, )


    @action(detail=False, methods=['get'], name='report_departments')
    def report_departments(self, request, *args, **kwargs):
        access_departments =request.user.profile.access.all()
        departments_names =[]
        for a in access_departments:
            departments_names.append(a.title)
        page_obj = get_page_obj(self.queryset.filter(title__in=departments_names), 5, request)
        service = Service.objects.first()
        return render(request, 'IC/report_departments.html',
                      {'page_obj': page_obj, 'list': self.queryset, 'service': service})

    def get_deps(self):
        access_departments = self.request.user.profile.access.all()
        return access_departments


# формы добвления/изменения отдела/группы
@api_view(['GET', 'POST'])
@parser_classes([MultiPartParser])
@permission_classes([IsCreator])
def forms_department(request, department_id=None):

    if department_id is None:
        department = None
        template = 'IC/add_object.html'
        context = {
            'url_name': 'add_department',
            'title': 'Добавление отдела/группы',
        }
    else:
        department = Department.objects.get(pk=department_id)
        template = 'IC/change_object.html'
        context = {
            'url_name': 'change_department',
            'title': 'Редактирование отдела/группы',
        }
    if request.method == 'POST':
        form = DepartmentForm(request.POST, request.FILES, instance=department)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = DepartmentForm(instance=department)
        context['form'] = form
    return render(request, template, context)


# Набор представлений показателей

class IndicatorViewSet(ModelViewSet):
    queryset = Indicator.objects.select_related(
        'department')
    #permission_classes = (IsAuthenticated,)
    serializer_class = IndicatorSerializer


    # представление для Расчётной формы
    @action(detail=False, methods=['get'], name='Settlement form')
    def settlement_form(self, request, *args, **kwargs):
        page_obj = get_page_obj(self.queryset.filter(department__in=request.user.profile.access.all()), 5, request)
        return render(request, 'IC/settlement_form.html',
                      {'page_obj': page_obj, 'list': self.queryset})

    # представление для просмотра подробностей показателя
    @action(detail=True, methods=['get'], name='Show Details')
    def show_details(self, request, pk):
        indicator = Indicator.objects.get(pk=pk)
        if indicator.department in request.user.profile.access.all():
            files = indicator.indicators_file_set.all()
            return render(request, 'IC/Indicator_detail.html',
                        { 'indicator': indicator, 'files': files})
        return HttpResponse(status=403)

    # представление для ввода данных

@api_view(['GET', 'POST'])
@parser_classes([MultiPartParser])
@permission_classes([IsCreator])
def data_input(request, indicator_id=None):
        page_obj = get_page_obj(Indicator.objects.filter(department__in=request.user.profile.access.all()), 6, request)
        context = { 'url_name': 'input2'}
        forms = []
        file_form = Indicators_fileForm(None)
        if request.method == 'POST':
            indicator_form = IndicatorForm(request.POST, instance=Indicator.objects.get(pk=indicator_id))
            file_form = Indicators_fileForm(request.POST, request.FILES)
            files = request.FILES.getlist('confirmation_document')
            if indicator_form.is_valid() and file_form.is_valid():
                indicator_instance = indicator_form.save(commit=False)
                indicator_instance.save()
                for f in files:
                    indicators_file_instance = Indicators_file(confirmation_document=f, indicator=indicator_instance)
                    indicators_file_instance.save()
                return redirect('input1')
        else:
            for i in page_obj:
                indicator_form = IndicatorForm(instance=i)
                forms.append(indicator_form)
            file_form = Indicators_fileForm(None)
            context['fform'] = file_form
            context['forms'] = forms
            context['page_obj'] = page_obj
        return render(request, 'IC/data.html', context)






# представление для удаления/добавления файлов подтверждения значения показателя
#@permission_required([IsCreator], raise_exception=True)
@api_view(['GET', 'POST'])
@parser_classes([MultiPartParser])
@permission_classes([IsCreator])
def change_files(request, indicator_id=None ):
    indicator = Indicator.objects.get(pk=indicator_id)
    if indicator.department not in request.user.profile.access.all():
         return HttpResponse(status=403)
    files = indicator.indicators_file_set.all()
    template = 'IC/change_files.html'
    context = {
        'url_name': 'change_files',
    }
    if request.method == 'POST':
        file_form = Indicators_fileForm(request.POST, request.FILES)
        files = request.FILES.getlist('confirmation_document')
        if file_form.is_valid():
            for f in files:
                indicators_file_instance = Indicators_file(confirmation_document=f, indicator=indicator)
                indicators_file_instance.save()
            return redirect('change_files', indicator_id)
        # indicator - show - details
    else:
        file_form = Indicators_fileForm(None)
        context['form'] = file_form
        context['indicator_id'] = indicator.id
        context['indicator_title'] = indicator.title
        context['files'] = files
    return render(request, template, context)


# формы добвления/изменения показателя
#@permission_required([IsCreator], raise_exception=True)
@api_view(['GET', 'POST'])
@parser_classes([MultiPartParser])
@permission_classes([IsCreator])
def forms_indicator(request, indicator_id=None):
    if indicator_id is None:
        indicator = None

        indicators_file = None
        template = 'IC/add_object.html'
        context = {
            'url_name': 'add_indicator',
            'title': 'Добавление показателя',
        }
    else:
        indicator = Indicator.objects.get(pk=indicator_id)
        template = 'IC/change_object.html'
        context = {
            'url_name': 'change_indicator',
            'title': 'Редактирование показателя',
        }
    if request.method == 'POST':
        indicator_form = IndicatorForm(request.POST, instance=indicator)
        file_form = Indicators_fileForm(request.POST, request.FILES)
        files = request.FILES.getlist('confirmation_document')
        if indicator_form.is_valid() and file_form.is_valid():
            indicator_instance = indicator_form.save(commit=False)
            indicator_instance.save()
            for f in files:
                indicators_file_instance = Indicators_file(confirmation_document=f, indicator=indicator_instance)
                indicators_file_instance.save()
            return redirect('home')
    else:
        indicator_form = IndicatorForm(instance=indicator)
        file_form = Indicators_fileForm(None)
        context['form'] = indicator_form
        context['fform'] = file_form
    return render(request, template, context)




# набор представлений для файлов подтверждения показателя

class Indicators_fileViewSet(ModelViewSet):
    queryset = Indicators_file.objects.select_related(
        'indicator')
    serializer_class = Indicators_fileSerializer
    permission_classes =()

#@permission_required([IsCreator], raise_exception=True)
@api_view(['GET', 'DELETE'])
@parser_classes([MultiPartParser])
@permission_classes([IsCreator])
def delete_file(request, id):
    try:
        file =  Indicators_file.objects.get(id=id)
        file.delete()
        return redirect('change_files', file.indicator.id)
    except  Indicators_file.DoesNotExist:
        return Response(status=status.HTTP_204_NO_CONTENT)



# набор представлений для критических сервисов/служб
class Critical_serviceViewSet(ModelViewSet):
    queryset = Critical_service.objects.all()
    serializer_class = Critical_serviceSerializer
    permission_classes =  (IsAuthenticated, IsAccess)

# формы добвления/изменения критических сервисов/служб
#@permission_required([IsCreator], raise_exception=True)
@api_view(['GET', 'POST'])
@permission_classes([IsCreator])
def forms_crirtical_service(request, critical_service_id=None):
    if critical_service_id is None:
        cs = None
        template = 'IC/add_object.html'
        context = {
            'url_name': 'add_critical_service',
            'title': 'Добавление критически важного сервиса/службы',
        }
    else:
        cs = Critical_service.objects.get(pk=critical_service_id)
        template = 'IC/change_object.html'
        context = {
            'url_name': 'change_critical_service',
            'title': 'Редактирование критически важного сервиса/службы',
        }
    if request.method == 'POST':
        form = Crtitical_serviceForm(request.POST, instance=cs)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = Crtitical_serviceForm(instance=cs)
        context['form'] = form
    return render(request, template, context)

#@permission_required([IsCreator], raise_exception=True)
@api_view(['GET', 'POST'])
@permission_classes([IsCreator])
def CS_data_input(request, CS_id=None):
    page_obj = get_page_obj(Critical_service.objects.all(), 6, request)
    context = {'url_name': 'CS_input2'}
    forms = []
    if request.method == 'POST':
        CS_form = Crtitical_serviceForm(request.POST, instance=Critical_service.objects.get(pk=CS_id))
        if CS_form.is_valid():
            indicator_instance = CS_form.save(commit=False)
            indicator_instance.save()
            return redirect('CS_input1')
        else:
            print(CS_form.errors)
    else:
        for i in page_obj:
            CS_form = Crtitical_serviceForm(instance=i)
            forms.append(CS_form)
        context['forms'] = forms
        context['page_obj'] = page_obj
    return render(request, 'IC/critical_services.html', context)


# Представление для регистрации пользователя
class RegisterUser(CreateView):
    form_class = RegisterUserForm
    template_name = 'IC/register.html'
    success_url = reverse_lazy('login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        return dict(list(context.items()))

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        return redirect('home')

# Представление для авторизации пользователя
class LoginUser(LoginView):
    form_class = LoginUserForm
    template_name = 'IC/login.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)

        return dict(list(context.items()))

    def get_success_url(self):
        return reverse_lazy('home')

# Функция выхода из аккаунта
def logout_user(request):
    logout(request)
    return redirect('login')



def print_report(request):

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    print_report_departments(request, workbook, 1)
    print_settlement_form(request, workbook, 2)
    print_report_CS(workbook, 3)
    workbook.save(response)
    return response


